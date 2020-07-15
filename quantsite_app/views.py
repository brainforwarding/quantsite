from django.shortcuts import render
from django.http import HttpResponse
from django.template import Context, loader
from rest_framework.decorators import api_view, permission_classes
from io import BytesIO
from openpyxl import Workbook
from django.core.mail import EmailMessage
from rest_framework.response import Response
from django.core.mail import EmailMessage
import base64
import openpyxl

@api_view(['GET'])
def index(request):
    template = loader.get_template("quantsite_app/index.html")
    return HttpResponse(template.render())


@api_view(['POST'])
def post(request):
    name = request.POST.get('name')
    email = request.POST.get('email')
    message = request.POST.get('message')
    filename = request.POST.get('f_name')
    xls_file = request.FILES.get('xls_file')
    wb = openpyxl.load_workbook(filename=BytesIO(xls_file.read()))
    sheet = wb.active
    cell = sheet.cell(row=sheet.max_row+1,column=1)
    cell.value = "I've been modified!"
    wb.save(xls_file.file)
    email_obj = EmailMessage(
    'Modified Excel File',
    f'Your name: {name} \n Your message: {message} \n Find the modified xls file attached!',
    'haziq.usman98@gmail.com',
    [email],
    headers={'Message-ID': 'foo'},
)

    email_obj.attach(filename, xls_file.file.getvalue(), 'application/vnd.ms-excel')
    email_obj.send()

    return Response(f'Email sent to: {email}')

